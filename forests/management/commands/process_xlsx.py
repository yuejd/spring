from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook
import re
from forests.models import Server, Employee, Team


class Command(BaseCommand):
    help = "import hosts info from the excel fiel"

    def process_excel(self):
        try:
            wb = load_workbook(self.filename)
        except:
            raise CommandError("failed to load the excel file")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows('A1:J' + str(sheet.max_row)):
                if row[9].value and row[1].value != "Reserved":
                    yield (
                        row[1].value,
                        row[7].value,
                        row[9].value
                        )

    def add_arguments(self, parser):
        parser.add_argument('filename', nargs='+', type=str)

    def handle(self, *args, **options):
        self.filename = options.get("filename")[0]
        t = Team.objects.get_or_create(name="x86")[0]
        for item in self.process_excel():
            (user, host_name, ip) = item
            ip_re = re.search(r"\d+(\.\d+){3}", ip)
            if not ip_re:
                print("failed to process:")
                print(item)
                continue
            else:
                ip = ip_re.group()
            if user:
                u = Employee.objects.get_or_create(
                    team=t,
                    name=user,
                    )[0]
                Server.objects.get_or_create(
                    ip_addr=ip,
                    host_name=host_name,
                    owner=u,
                    team=t,
                    username="root",
                    password="#1Danger0us",
                    )
            else:
                Server.objects.get_or_create(
                    ip_addr=ip,
                    host_name=host_name,
                    team=t,
                    username="root",
                    password="#1Danger0us",
                    )
